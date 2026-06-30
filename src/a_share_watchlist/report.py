from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .compliance import assert_no_banned_terms
from .config import ScreeningConfig


def write_excel_report(report: pd.DataFrame, output_path: str | Path, config: ScreeningConfig) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    assert_no_banned_terms(report.to_numpy().ravel())

    params = pd.DataFrame(
        [
            ["min_amount", config.min_amount],
            ["min_listing_days", config.min_listing_days],
            ["min_turnover", config.min_turnover],
            ["max_turnover", config.max_turnover],
            ["amount_spike_multiple", config.amount_spike_multiple],
            ["max_5d_gain", config.max_5d_gain],
            ["ma_window", config.ma_window],
            ["as_of_date", config.as_of_date or "数据中最新交易日"],
        ],
        columns=["参数", "值"],
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        report.to_excel(writer, index=False, sheet_name="观察名单")
        params.to_excel(writer, index=False, sheet_name="参数")

        workbook = writer.book
        _format_sheet(workbook["观察名单"])
        _format_sheet(workbook["参数"])

    return output


def _format_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True, color="1F2933")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 14,
        "B": 18,
        "C": 18,
        "D": 46,
        "E": 52,
        "F": 14,
    }
    for col_idx in range(1, sheet.max_column + 1):
        letter = get_column_letter(col_idx)
        sheet.column_dimensions[letter].width = widths.get(letter, 22)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
