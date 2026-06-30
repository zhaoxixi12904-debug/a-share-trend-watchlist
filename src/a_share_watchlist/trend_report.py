from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .compliance import assert_no_banned_terms


def write_trend_report(
    thesis: pd.DataFrame,
    watchlist: pd.DataFrame,
    trend_snapshot: pd.DataFrame,
    output_path: str | Path,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    assert_no_banned_terms(thesis.to_numpy().ravel())
    assert_no_banned_terms(watchlist.to_numpy().ravel())
    assert_no_banned_terms(trend_snapshot.to_numpy().ravel())

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        thesis.to_excel(writer, index=False, sheet_name="行业判断")
        trend_snapshot.to_excel(writer, index=False, sheet_name="走势概览")
        watchlist.to_excel(writer, index=False, sheet_name="观察名单")
        _format_sheet(writer.book["走势概览"], {"A": 14, "B": 18, "C": 20, "D": 14, "E": 14, "F": 14, "G": 16, "H": 18, "I": 12})
        _format_sheet(writer.book["行业判断"], {"A": 18, "B": 18, "C": 50, "D": 48, "E": 42})
        _format_sheet(writer.book["观察名单"], {"A": 14, "B": 18, "C": 20, "D": 52, "E": 58, "F": 14})
        _add_visuals(writer.book["走势概览"])

    return output


def _format_sheet(sheet, widths: dict[str, int]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True, color="1F2933")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx in range(1, sheet.max_column + 1):
        letter = get_column_letter(col_idx)
        sheet.column_dimensions[letter].width = widths.get(letter, 22)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _add_visuals(sheet) -> None:
    from openpyxl.chart import BarChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule

    if sheet.max_row < 2:
        return

    # 5/20/60 日涨幅用红白绿三色刻度，便于一眼看强弱。
    for column in ("D", "E", "F"):
        sheet.conditional_formatting.add(
            f"{column}2:{column}{sheet.max_row}",
            ColorScaleRule(
                start_type="num",
                start_value=-0.2,
                start_color="F4CCCC",
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFF",
                end_type="num",
                end_value=0.2,
                end_color="D9EAD3",
            ),
        )

    # 成交额放大倍数用数据条。
    sheet.conditional_formatting.add(
        f"G2:G{sheet.max_row}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=3, color="6FA8DC"),
    )

    # 趋势图标：上行、走平、转弱。
    sheet.conditional_formatting.add(
        f"I2:I{sheet.max_row}",
        IconSetRule("3Arrows", "num", [0, 50, 80], showValue=True),
    )

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "近20日涨幅概览"
    chart.y_axis.title = "股票"
    chart.x_axis.title = "近20日涨幅"
    data = Reference(sheet, min_col=5, min_row=1, max_row=min(sheet.max_row, 21))
    cats = Reference(sheet, min_col=2, min_row=2, max_row=min(sheet.max_row, 21))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 22
    sheet.add_chart(chart, "K2")
